"""스코어링 엔진 — 정량 지표 자동 계산 + 정성 루브릭 병합.

사용:
  python score.py --group skincare         # 지표 계산 + 루브릭 병합 + xlsx
  python score.py --group all --no-xlsx     # 콘솔 출력만

산출:
  out/scorecard_{group}.xlsx — 샵 × 전 지표 매트릭스
    · 자사 행 하이라이트
    · 헤드라인 지표(리뷰볼륨·추정매출·디자인총점)에 '1위 대비 갭' 컬럼

루브릭 입력 우선순위: rubric_{group}.csv  >  design_scores 테이블.
"""
import sys
from pathlib import Path

import pandas as pd

from db import connect
from estimator import estimate, summary
from shops_groups import group_label, own_shop_ids, resolve_group

OUT = Path(__file__).parent / "out"

# 디자인 루브릭 가중치 (합 1.00)
RUBRIC_WEIGHTS = {
    "thumbnail": 0.25,
    "shop_main": 0.20,
    "localization": 0.20,
    "promo_design": 0.20,
    "consistency": 0.15,
}

# xlsx 한글 헤더
COLUMN_LABELS = {
    "shop_id": "샵", "role": "구분", "sku_count": "상품수",
    "price_min": "최저가", "price_median": "중앙가", "price_max": "최고가",
    "review_volume": "리뷰볼륨", "top5_power": "Top5파워",
    "set_ratio": "세트비중", "jp_limited_count": "일본한정수",
    "follower_count": "팔로워", "satisfaction_pct": "만족도%", "seller_grade": "등급",
    "promo_intensity": "프로모션강도", "est_revenue_30d": "추정매출(30일)",
    "design_total": "디자인총점", "design_rank": "디자인순위",
    "gap_review_volume": "리뷰볼륨_갭", "gap_est_revenue_30d": "추정매출_갭",
    "gap_design_total": "디자인총점_갭",
}


def _arg(flag: str, default=None):
    return sys.argv[sys.argv.index(flag) + 1] if flag in sys.argv else default


# ── 정량 지표 ──────────────────────────────────────────────────────────────
def _latest_products(conn, shop_ids: list[str]) -> pd.DataFrame:
    """그룹 샵들의 상품별 최신 스냅샷 1행씩."""
    q = f"""SELECT p.shop_id, s.goods_code, s.snap_date, s.price, s.review_count,
                   s.is_set, s.is_jp_limited, s.discount_pct
            FROM snapshots s JOIN products p USING(goods_code)
            WHERE p.shop_id IN ({",".join("?" * len(shop_ids))})"""
    df = pd.read_sql(q, conn, params=shop_ids)
    if df.empty:
        return df
    df = df.sort_values("snap_date").groupby("goods_code", as_index=False).tail(1)
    return df


def quant_metrics(conn, shop_ids: list[str]) -> pd.DataFrame:
    df = _latest_products(conn, shop_ids)
    if df.empty:
        return pd.DataFrame(columns=["shop_id"])

    rows = []
    for sid, g in df.groupby("shop_id"):   # 수동 루프(pandas 2.0~3.0 호환)
        rv = g["review_count"].fillna(0)
        total_rv = rv.sum()
        top5 = rv.nlargest(5).sum()
        rows.append({
            "shop_id": sid,
            "sku_count": len(g),
            "price_min": g["price"].min(),
            "price_median": g["price"].median(),
            "price_max": g["price"].max(),
            "review_volume": int(total_rv),
            "top5_power": round(top5 / total_rv, 3) if total_rv else None,
            "set_ratio": round(g["is_set"].fillna(0).mean(), 3),
            "jp_limited_count": int(g["is_jp_limited"].fillna(0).sum()),
        })
    return pd.DataFrame(rows)


def shop_level(conn, shop_ids: list[str]) -> pd.DataFrame:
    """샵 스냅샷 최신 1행 + 프로모션 원점수(coupon+timesale)."""
    q = f"""SELECT shop_id, snap_date, seller_grade, satisfaction_pct, follower_count,
                   coupon_count, timesale_count
            FROM shop_snapshots
            WHERE shop_id IN ({",".join("?" * len(shop_ids))})"""
    df = pd.read_sql(q, conn, params=shop_ids)
    if df.empty:
        return pd.DataFrame(columns=["shop_id"])
    df = df.sort_values("snap_date").groupby("shop_id", as_index=False).tail(1)
    df["promo_raw"] = df["coupon_count"].fillna(0) + df["timesale_count"].fillna(0)
    return df[["shop_id", "seller_grade", "satisfaction_pct", "follower_count", "promo_raw"]]


def est_revenue(shop_ids: list[str]) -> pd.DataFrame:
    """estimator 재사용 — 최근 30일 샵별 추정 매출 합 (2일+ 스냅샷 필요)."""
    df = estimate()
    if df.empty or not df["new_reviews"].notna().any():
        return pd.DataFrame(columns=["shop_id", "est_revenue_30d"])
    s = summary(df)
    s = s[s["shop_id"].isin(shop_ids)]
    out = (s.groupby("shop_id", as_index=False)["est_revenue_jpy"].sum()
           .rename(columns={"est_revenue_jpy": "est_revenue_30d"}))
    return out


# ── 디자인 루브릭 ───────────────────────────────────────────────────────────
def design_rubric(conn, group: str, shop_ids: list[str]) -> pd.DataFrame:
    """rubric_{group}.csv 우선, 없으면 design_scores 테이블. 가중 총점 + 순위."""
    csv = Path(__file__).parent / f"rubric_{group}.csv"
    cols = list(RUBRIC_WEIGHTS)
    if csv.exists():
        df = pd.read_csv(csv)
    else:
        df = pd.read_sql(
            f"""SELECT shop_id, {", ".join(cols)} FROM design_scores
                WHERE shop_id IN ({",".join("?" * len(shop_ids))})""",
            conn, params=shop_ids)
    if df.empty:
        return pd.DataFrame(columns=["shop_id", "design_total", "design_rank"])

    # 채점값은 숫자 강제변환(CSV 오염·공란 방어)
    for c in cols:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    # 2인 교차 채점 등 다중 행은 샵별 평균 (편향 완화)
    agg = df.groupby("shop_id", as_index=False)[cols].mean()
    agg["design_total"] = sum(agg[c] * w for c, w in RUBRIC_WEIGHTS.items()).round(3)
    agg["design_rank"] = agg["design_total"].rank(ascending=False, method="min").astype(int)
    return agg[["shop_id", *cols, "design_total", "design_rank"]]


# ── 스코어카드 조립 ─────────────────────────────────────────────────────────
def build_scorecard(group: str) -> pd.DataFrame:
    shops = resolve_group(group)
    shop_ids = [s["shop_id"] for s in shops]
    role = {s["shop_id"]: s["role"] for s in shops}

    conn = connect()
    base = pd.DataFrame({"shop_id": shop_ids})
    for part in (quant_metrics(conn, shop_ids), shop_level(conn, shop_ids),
                 est_revenue(shop_ids), design_rubric(conn, group, shop_ids)):
        if not part.empty:
            base = base.merge(part, on="shop_id", how="left")

    base.insert(1, "role", base["shop_id"].map(role))

    # 프로모션 강도: 그룹 내 min-max 0~100 표준화
    if "promo_raw" in base:
        lo, hi = base["promo_raw"].min(), base["promo_raw"].max()
        rng = (hi - lo) or 1
        base["promo_intensity"] = ((base["promo_raw"] - lo) / rng * 100).round(1)
        base = base.drop(columns=["promo_raw"])

    # 1위 대비 갭 컬럼 (higher-is-better 헤드라인 지표)
    for metric in ("review_volume", "est_revenue_30d", "design_total"):
        if metric in base and base[metric].notna().any():
            base[f"gap_{metric}"] = (base[metric] - base[metric].max()).round(2)

    # 정렬: 리뷰볼륨 내림차순 (판매력 대리지표)
    if "review_volume" in base:
        base = base.sort_values("review_volume", ascending=False, na_position="last")
    return base.reset_index(drop=True)


# ── xlsx 출력 ──────────────────────────────────────────────────────────────
def write_xlsx(df: pd.DataFrame, group: str, path: Path) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    own = own_shop_ids(group)
    out = df.rename(columns=COLUMN_LABELS)
    OUT.mkdir(exist_ok=True)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        out.to_excel(writer, index=False, sheet_name=group[:31] or "scorecard")
        ws = writer.sheets[group[:31] or "scorecard"]

        header_fill = PatternFill("solid", fgColor="2F5597")
        own_fill = PatternFill("solid", fgColor="FFF2CC")
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = header_fill
            c.alignment = Alignment(horizontal="center", vertical="center")

        # 자사 행 하이라이트 (1열은 shop_id)
        for r in range(2, ws.max_row + 1):
            if ws.cell(r, 1).value in own:
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = own_fill

        # 열 너비 자동(상한)
        for col in ws.columns:
            width = max((len(str(c.value)) for c in col if c.value is not None), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(width + 3, 22)
        ws.freeze_panes = "B2"


def main():
    group = _arg("--group", "skincare")
    df = build_scorecard(group)
    if df.empty or df.drop(columns=["shop_id", "role"]).isna().all().all():
        print(f"[{group}] 스코어 데이터 없음 — 먼저 collect.py 로 수집하세요.")
        return

    pd.set_option("display.max_columns", None, "display.width", 200)
    print(f"=== 스코어카드: {group} ({group_label(group)}) ===")
    print(df.to_string(index=False))

    if "--no-xlsx" not in sys.argv:
        path = OUT / f"scorecard_{group}.xlsx"
        write_xlsx(df, group, path)
        print(f"\n스코어카드 저장: {path}")


if __name__ == "__main__":
    main()
